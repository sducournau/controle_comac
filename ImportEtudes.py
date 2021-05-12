 #list_fields_etude = ['ETUDE','Path','Comac','Pdf','Commune','Rue','Supports','Lines']
import os
from os import listdir
from os.path import isfile, join
import xml.etree.ElementTree as ET
import csv
import json
import re
import sys
from pathlib import Path
import pyproj
from shapely import geometry
from shutil import copyfile
import openpyxl
from . import utils as utils
from .utils import *
import math
GESTIONNAIRE = utils.GESTIONNAIRE
# from qgis.PyQt.QtGui import *
# from qgis.PyQt.QtWidgets import *
# from qgis.utils import iface

class list_etudes_obj:



    def __init__(self, inputs, output, plugin_dir, layer_name, repare_pcm=False, option_boitier_fibre=False):

        self.list_fields_export_comac_supports = ['NRO','PMZ','Etude','etat','troncon', 'numero','commentair']
        self.list_fields_export_stats = ["NRO","Numéro PMZ GEOFIBRE",'Commune',"INSEE","Nom d'affaire E-PLAN",'Adresse','N° Plan','Nb de Support Enedis','Nb de Support D3','Réseau projetés (m)',
                                   'Longueur à facturer ENEDIS (formule)','Nb de supports à créer pour portée','Nb de supports à créer pour renfort',"Indice","Type","Type_Calculé",'Longueur à facturer ENEDIS (absolu)','path','pcm','pdf']
        self.list_fields_export_capft_supports = ['NRO','PMZ','Etude', 'nom_norma', 'controle_visuel','controle_calcul','etat','ban_vert','modele', 'sup_remp', 'numero','com_etat','propriet','h_traverse', 'commentair']
        self.inputs = []
        for input in inputs:
            self.inputs.append('\\\\?\\' + os.path.normpath(input))

        self.nb_appuis_capft = 0
        self.nb_appuis_comac = 0
        self.output = output
        self.list_pcm = []
        self.list_C6 = []
        self.list_etudes_comac = []
        self.list_etudes_capft = []
        self.data = {'comac' : self.list_etudes_comac, 'capft':self.list_etudes_capft }
        self.plugin_dir = plugin_dir
        self.layer_name = layer_name
        print(self.inputs)
        with open(self.plugin_dir + os.sep + 'list_supports.txt', newline='', encoding='cp1252') as csv_trad_supports:
            reader = csv.DictReader(csv_trad_supports, delimiter='\t')
            dict_list = []
            for line in reader:
                dict_list.append(line)

        self.type_supports = dict_list
        self.repare_pcm = repare_pcm
        self.option_boitier_fibre = option_boitier_fibre

        if len(self.list_pcm) == 0 or len(self.list_C6) == 0:
            self.get_list_etudes()


    def remove_duplicate(self, type_etude):

        name_list = []
        for i, etude in enumerate(self.data[type_etude]):
            if etude['Etude'] in name_list:
                print("Etude duppliquée",etude['Etude'])
                self.data[type_etude].pop(i-1)
                if type_etude == 'comac':
                    self.nb_appuis_comac -= len(etude['Supports'])
                if type_etude == 'capft':
                    self.nb_appuis_capft -= len(etude['Supports'])
            if etude['Etude'] not in name_list:
                name_list.append(etude['Etude'])



    def calcul(self, layer_name):
        if layer_name == 'appuis_comac':
            self.iterate_over_comac_etudes()
            self.remove_duplicate('comac')
            print(str(len(self.data['comac'])) + ' études comac ont été calculées')
        if layer_name == 'appuis_capft':
            self.iterate_over_capft_etudes()
            self.remove_duplicate('capft')
            # self.remove_duplicate_capft()
            print(str(len(self.data['capft'])) + ' C6 globaux ft ont été calculées')
        if layer_name == 'All':
            self.iterate_over_comac_etudes()
            self.iterate_over_capft_etudes()
            self.remove_duplicate('comac')
            self.remove_duplicate('capft')
            # self.remove_duplicate_capft()
            print(str(len(self.data['comac'])) + ' études comac ont été calculées')
            print(str(len(self.data['capft'])) + ' études capft ont été calculées')



    def export(self, options_export):
        self.exportEtudeItem(self.output, self.data, self.list_fields_export_comac_supports, self.list_fields_export_stats, self.list_pcm,self.list_C6, self.list_fields_export_capft_supports, options_export, self.layer_name)


    def get_list_etudes(self):

        i = 0

        for input in self.inputs:
            if self.layer_name == 'appuis_comac':
                for subdir, dirs, files in os.walk(input):

                    i, pcm = self.get_pcm(subdir, i)
                    if pcm:
                        self.list_pcm += pcm

            elif self.layer_name == 'appuis_capft':
                C6 = self.calculate_C6(input)
                if C6:
                    self.list_C6 += C6


    def calculate_C6(self, input):
        sheet = None
        list_C6 = []

        # try:
        name_file = Path(input)
        name_file = name_file.name

        pmz =re.search('((PM){1}Z{0,1}[ _-]{0,1}){1}[0-9]*', input)[0] if re.search('((PM){1}Z{0,1}[ _-]{0,1}){1}[0-9]*', input) else 'Autres'

        pmz = re.sub('((PM){1}Z{0,1}[ _-]{0,1}){1}[0-9]*', '', pmz)

        nro = re.sub('[()]*', '', re.search('(?<=NRO)([^\\\\]*)(?=\\\\)' ,input)[0])[-3:] if re.search('(?<=NRO)([^\\\\]*)(?=\\\\)' ,input) else None
        if nro is None:
            nro = re.search('(?<=NRO)([A-Z0-9]*)' ,input)[0] if re.search('(?<=NRO)([A-Z0-9]*)' ,input) else None

        wb_obj = openpyxl.load_workbook(input,data_only=True)
        for y, sheetname in enumerate(wb_obj.sheetnames):

            if 'Export' in sheetname:
                sheet = wb_obj.worksheets[y]

        list_C6.append({'path': input, 'xls' : sheet, 'directory' : os.path.dirname(input), 'modified':os.path.getmtime(input), 'pmz': pmz, 'nro' : nro})

        # except:
        #     print(input)

        return list_C6




    def get_pcm(self, directory, i):



        # duplicate_list_pcms = []
        i, list_pcms = self.calculate_pcm(directory, i)
        #
        # duplicate_list_pcms = [x for x in list_pcms if list_pcms.count(x) > 1]
        # for duplicate_pcms in duplicate_list_pcms:
        #     sorted_list_pcms = []
        #     sorted_list_pcms = sorted(list_pcms, reverse=True, key=lambda k: (k['version'], k['modified']))
        #
        return i, list_pcms if list_pcms else None



    def calculate_pcm(self, directory, i):

        list_pcms = []
        onlyfiles = [f for f in listdir(directory) if isfile(join(directory, f))]

        for file in onlyfiles:


            filepath = directory + os.sep + file


            if filepath.upper().endswith(".PCM") and 'ARCHIVE' not in filepath.upper():
                try:
                    tree = ET.parse(filepath)
                    root = tree.getroot()
                    name_file = Path(filepath)
                    name_file = name_file.name

                    pmz = re.search('((PM){1}Z{0,1}[ _-]{0,1}){1}[0-9]*', name_file)[0] if re.search('((PM){1}Z{0,1}[ _-]{0,1}){1}[0-9]*', name_file) else None
                    if pmz is None:
                        pmz =re.search('((PM){1}Z{0,1}[ _-]{0,1}){1}[0-9]*', filepath)[0] if re.search('((PM){1}Z{0,1}[ _-]{0,1}){1}[0-9]*', filepath) else 'Autres'

                    pmz = re.sub('(PM{1}Z{0,1}[ _-]{0,1})', '', pmz)


                    version = re.search('(V{1})([0-3]{1})', name_file)[0] if re.search('(V{1})([0-3]{1})', name_file) else 'V1'

                    nro = re.sub('[()]*', '', re.search('(?<=NRO)([^\\\\]*)(?=\\\\)' ,filepath)[0])[-3:] if re.search('(?<=NRO)([^\\\\]*)(?=\\\\)' ,filepath) else None
                    if nro is None:
                        nro = re.search('(?<=NRO)([A-Z0-9]*)' ,filepath)[0] if re.search('(?<=NRO)([A-Z0-9]*)' ,filepath) else None

                    list_pcms.append({'id': i, 'path': filepath, 'xml' : root, 'directory' : directory, 'modified':os.path.getmtime(filepath), 'version' : version, 'pmz': pmz, 'nro' : nro})
                    i+=1

                except:
                    print('Error PCM : ', filepath)



        return i, list_pcms



    def iterate_over_comac_etudes(self):

        print(len(self.list_pcm))

        for y, pcm in enumerate(self.list_pcm):

            self.buildEtudeComacItem(pcm, self.data['comac'], self.list_fields_export_comac_supports, self.type_supports, self.repare_pcm, self.option_boitier_fibre)
            self.nb_appuis_comac += len(self.list_etudes_comac[y]['Supports'])


    def iterate_over_capft_etudes(self):

        for y, C6 in enumerate(self.list_C6):

            self.buildEtudeFtItem(C6, self.data['capft'], self.list_fields_export_capft_supports)

            self.nb_appuis_capft += len(self.list_etudes_capft[y]['Supports'])




    def update_etude_comac(self, etude, supports_to_update):

        tree = ET.parse(etude['Pcm'])
        print('start')
        for nom_appui in supports_to_update:
            for support in tree.findall('Supports//Support'):
                for field in support:
                    if field.tag == 'Nom':
                        if field.text == nom_appui:
                            for field_ in support:
                                if field_.tag == 'optBoitierFibre':
                                    field_.text = str(1)

        tree.write(etude['Pcm'])


        for feat in self.list_etudes_comac:
            if etude['id'] == feat['id']:
                self.list_etudes_comac.remove(feat)


        for pcm in self.list_pcm:

            if pcm['id'] == etude['id']:
                tree = ET.parse(etude['Pcm'])
                root = tree.getroot()
                pcm['xml'] = root
                self.buildEtudeComacItem(pcm, self.data, self.list_fields_export_comac_supports, self.type_supports)




    def update_pcms_appuis(self, layer):

        dict_to_replace ={}

        i = 0
        for y, pcm in enumerate(self.list_pcm):

            tree = ET.parse(pcm['path'])
            root = tree.getroot()
            i += 1


            for k,  item in enumerate(layer.getFeatures()):
                old_name =''
                new_name =''

                if item['Etude'] in pcm['path']:

                    for support in root.findall('Supports//Support'):
                        for field in support:

                            if field.text == item['Nom']:
                                old_name = field.text
                                if item['propriet'] == GESTIONNAIRE:
                                    field.text = GESTIONNAIRE + re.sub('[A-Z]*','', item['Nom'])
                                if item['Nature'] in ('BE', 'BO','ME'):
                                    if item['join_numero'] != None:
                                        field.text = re.sub(item['numero'],re.sub('/[0-9]*', '', item['join_numero']), item['Nom'])
                                dict_to_replace[item['Nom']] = field.text
                                new_name = field.text





                    for support in root.findall('Portees//Portee'):
                        for field in support:
                            if field.text == item['Nom']:
                                field.text = new_name




                    for ligne in root.findall('LignesBT//LigneBT'):
                        for field in ligne.findall('Supports//Support'):
                            if field.text == item['Nom']:
                                field.text = new_name




                    for ligne in root.findall('LignesTCF//LigneTCF'):
                        for field in ligne.findall('Supports//Support'):
                            if field.text == item['Nom']:
                                field.text = new_name

                    try:

                        if item['propriet'] == GESTIONNAIRE or item['Nature'] in ('BE', 'BO','ME') and old_name != new_name:


                            for subdir, dirs, files in os.walk(os.path.dirname(pcm['path'])):
                                for file in files:
                                    filepath = subdir + os.sep + file
                                    if filepath.endswith(".jpg"):
                                        space = False
                                        new_filename = re.sub(old_name, new_name,  filepath)
                                        if file.find(' ') > 0: space = True

                                        if new_filename != filepath:
                                            os.rename(filepath, new_filename)
                                        elif space:
                                            new_filename = re.sub(item['numero'], new_name,  file)
                                            if new_filename:
                                                os.rename(filepath, subdir + os.sep + new_filename)

                                        print(new_filename)
                    except:
                        pass


            tree.write(pcm['path'], encoding="iso-8859-1",xml_declaration=True)


            # try:
            #     self.update_pdfs_appuis(dict_to_replace, pcm['path'].replace('.PCM', '.pdf'))
            # except:
            #     pass


        return i




    def __str__(self):
            return json.dumps(self.data, sort_keys=True, indent=4)



    def __len__(self):
            return len(self.data)




    class exportEtudeItem:


        def __init__(self,output, data, list_fields_export_comac_supports, list_fields_export_stats, list_pcms, list_C6, list_fields_export_capft_supports, options_export, layer_name):
            self.list_fields_export_comac_supports = list_fields_export_comac_supports
            self.list_fields_export_capft_supports = list_fields_export_capft_supports
            self.list_fields_export_stats = list_fields_export_stats
            self.list_etudes_comac = data['comac']
            self.list_etudes_capft = data['capft']
            self.list_pcms = list_pcms
            self.list_C6 = list_C6
            self.output_path = output
            self.output_directory = 'EXPORT_COMAC'
            self.options_export = options_export
            self.layer_name = layer_name
            self.encoding = 'iso-8859-1'

            if not os.path.isdir(self.output_path + os.sep + self.output_directory):
                os.mkdir(self.output_path + os.sep + self.output_directory)






            self.output = self.output_path + os.sep + self.output_directory + os.sep

            if self.layer_name == 'appuis_comac':
                print(self.layer_name)
                self.export_data_to_json()
                self.export_stats_to_csv()
                self.export_supports_comac_to_csv()
                self.export_portee_comac_to_json()

            if self.layer_name == 'appuis_capft':
                print(self.layer_name)
                self.export_supports_capft_to_csv()

            if self.options_export[0] == 2 and len(self.list_pcms) > 0: self.copy_pcms()


        def export_all(self):
            self.export_data_to_json()
            self.export_stats_to_csv()
            self.export_supports_comac_to_csv()
            self.export_supports_capft_to_csv()
            self.export_portee_comac_to_json()

        def copy_pcms(self):

            for pcm in self.list_pcms:
                name_pcm = Path(pcm['path'])
                name_pcm = name_pcm.name
                dir_nro = ''



                dir_NRO = 'NRO_' + pcm['nro'] if pcm['nro'] is not None else 'Autres'
                dir_PMZ = 'PMZ_' + pcm['pmz'] if pcm['pmz'] is not None else 'Autres'

                if not os.path.isdir(self.output + os.sep + dir_NRO):
                    os.mkdir(self.output + os.sep + dir_NRO)

                if not os.path.isdir(self.output + os.sep + dir_NRO + os.sep + dir_PMZ):
                    os.mkdir(self.output + os.sep + dir_NRO + os.sep + dir_PMZ)


                try:
                    if self.list_etudes_comac[pcm['id']]['Pdf'] is not None:
                        name_pdf = Path(self.list_etudes_comac[pcm['id']]['Pdf'])
                        name_pdf = name_pdf.name
                        copyfile('\\\\?\\' + self.list_etudes_comac[pcm['id']]['Pdf'], self.output + os.sep + dir_NRO + os.sep + dir_PMZ + os.sep + name_pdf)
                except:
                    print(pcm['id'])

                copyfile(pcm['path'], self.output + os.sep + dir_NRO + os.sep + dir_PMZ + os.sep + name_pcm)





            for C6 in self.list_C6:

                name_C6 = Path(C6['path'])
                name_C6 = name_C6.name

                dir_NRO = 'NRO_' + C6['nro'] if C6['nro'] is not None else 'Autres'
                dir_PMZ = 'PMZ_' + C6['pmz'] if C6['pmz'] is not None else 'Autres'

                if not os.path.isdir(self.output + os.sep + dir_NRO):
                    os.mkdir(self.output + os.sep + dir_NRO)

                if not os.path.isdir(self.output + os.sep + dir_NRO + os.sep + dir_PMZ):
                    os.mkdir(self.output + os.sep + dir_NRO + os.sep + dir_PMZ)

                copyfile(C6['path'], self.output + os.sep + dir_NRO + os.sep + dir_PMZ + os.sep + name_C6)









        def export_data_to_json(self):
            with open(self.output + 'export_data_comac.json', 'w') as outfile:
                json.dump(self.list_etudes_comac, outfile)

        def export_portee_comac_to_json(self):
            with open(self.output + 'export_portee_appuis_comac.json', 'w') as outfile:
                lines = {}
                for etude in self.list_etudes_comac:
                    lines[etude['Etude']] = etude['Lignes']

                json.dump(lines, outfile)



        def export_stats_to_csv(self):

            with open(self.output + 'export_tableau_suivi_enedis.csv', 'w', newline='', encoding=self.encoding) as outfile:


                writer = csv.DictWriter(outfile, fieldnames=self.list_fields_export_stats,delimiter=';')
                writer.writeheader()
                for etude in self.list_etudes_comac:
                    row = {}
                    num = re.sub('[A-Za-z]*', '', etude['Numero'])

                    i = 0
                    row[self.list_fields_export_stats[i]] = etude['NRO']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['PMZ']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Commune']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Insee']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Etude']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Rue']
                    i += 1
                    row[self.list_fields_export_stats[i]] = int(num) if num != '-' else num
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Infos']['Nombre_D2']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Infos']['Nombre_D3']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Infos']['Longueur_totale_pose']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Infos']['Longueur_totale_facture']['Formule']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Infos']['Nombre_ATHD_Portée']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Infos']['Nombre_ATHD_Renfort']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Version']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Type']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Infos']['Type_calculé']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Infos']['Longueur_totale_facture']['Absolue']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Path']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Pcm']
                    i += 1
                    row[self.list_fields_export_stats[i]] = etude['Pdf']
                    i += 1


                    writer.writerow(row)


        def export_supports_comac_to_csv(self):


            with open(self.output + 'export_appuis_comac.csv', 'w', newline='', encoding=self.encoding) as outfile:


                writer = csv.DictWriter(outfile, fieldnames=self.list_fields_export_comac_supports,delimiter=';')
                writer.writeheader()

                for etude in self.list_etudes_comac:
                    for support in etude['Supports']:
                        support['NRO'] = etude['NRO']
                        support['PMZ'] = etude['PMZ']
                        support['Etude'] = etude['Etude']
                        writer.writerow(support)


        def export_supports_capft_to_csv(self):


            with open(self.output + 'export_appuis_capft.csv', 'w', newline='', encoding=self.encoding) as outfile:


                writer = csv.DictWriter(outfile, fieldnames=self.list_fields_export_capft_supports,delimiter=';')
                writer.writeheader()

                for etude in self.list_etudes_capft:
                    for support in etude['Supports']:
                        writer.writerow(support)




    class buildEtudeFtItem:

        def __init__(self, C6, list_etudes, list_fields):


            self.list_fields_export_capft_supports = list_fields
            self.list_etudes_capft = list_etudes
            self.C6 = C6
            self.xls =  self.C6['xls']
            self.directory = self.C6['directory']
            self.name_1 = Path(self.C6['directory'])
            self.name_1 = self.name_1.name
            self.name_2 = Path(self.C6['path'])
            self.name_2 = self.name_2.name


            self.create_etude()

        def create_etude(self):

            def dms2dd(degrees, minutes, seconds, direction):
                dd = float(degrees) + float(minutes)/60 + float(seconds)/(60*60);
                if direction == 'E' or direction == 'N':
                    dd *= -1
                return dd;

            def dd2dms(deg):
                d = int(deg)
                md = abs(deg - d) * 60
                m = int(md)
                sd = (md - m) * 60
                return [d, m, sd]

            def parse_dms(dms):
                parts = re.split('[^\d\w]+', dms)
                lat = dms2dd(parts[0], parts[1], parts[2], parts[3])

                return (lat)

            def decdeg2dms(dd):
                negative = float(dd) < 0
                dd = abs(float(dd))
                minutes,seconds = divmod(dd*3600,60)
                degrees,minutes = divmod(minutes,60)
                if negative:
                    if degrees > 0:
                        degrees = -degrees
                    elif minutes > 0:
                        minutes = -minutes
                    else:
                        seconds = -seconds
                return negative, degrees,minutes,str(seconds)

            list_supports = []

            list_controle = ['Contrôle visuel OK','Contrôle verticalité OK et absence étiquette orange','Contrôle flamblement OK','Respect voisinage réseau électrique','Contrôle pointe carrée OK','Contrôle secousses OK','Contrôle percussion OK','Absence étiquette jaune','Appui utilisable en l état']

            sheet = self.xls

            etude_name = str(sheet.cell(column=3, row=3).value) + '|' + str(sheet.cell(column=5, row=3).value) + '|' + str(sheet.cell(column=7, row=3).value) + '|' + os.path.basename(self.C6['path']).upper()


            for col in range(1, sheet.max_column):
                if sheet.cell(column=col, row=8).value:
                    if sheet.cell(column=col, row=8).value.replace('\n', ' ').replace('\'', ' ').strip() not in self.list_fields_export_capft_supports:
                        self.list_fields_export_capft_supports.append(sheet.cell(column=col, row=8).value.replace('\n', ' ').replace('\'', ' ').strip())

            for row in range(9, sheet.max_row):
                support = {}
                support['com_etat'] = ''
                support['controle_visuel'] = 'nok'
                support['propriet'] = 'ORANGE'
                support['Etude'] = etude_name
                support['NRO'] = self.C6['nro']
                support['PMZ'] = self.C6['pmz']

                if sheet.cell(column=1, row=row).value is not None:

                    for col in range(1, sheet.max_column):

                        if self.list_fields_export_capft_supports[col+14] == 'N° appui':

                            val_str = ''
                            val_list = list(str(sheet.cell(column=col, row=row).value))

                            i = 0
                            while i < len(val_list):
                                print(val_list[i])
                                if val_list[i] == '0':
                                    val_list.pop(i)
                                    i += 1
                                else:
                                    break

                            for char in val_list:
                                val_str += char
                            print(val_str)
                            support['numero'] = str(val_str)


                        if self.list_fields_export_capft_supports[col+14] in ['Latitude (WGS84)', 'Longitude (WGS84)']:
                            support[self.list_fields_export_capft_supports[col+14]] = str(sheet.cell(column=col, row=row).value).replace(' ', '').replace(u"\u2018", "'").replace(u"\u2019", "'")

                        if self.list_fields_export_capft_supports[col+14] == 'Contrôle verticalité OK et absence étiquette orange':
                            support[self.list_fields_export_capft_supports[col+14]] =  sheet.cell(column=col, row=row).value
                            if sheet.cell(column=col, row=row).value == 'Non':
                                support['com_etat'] = 'Recalage'



                        if self.list_fields_export_capft_supports[col+14] == 'Effort disponible après ajout câble':
                            support[self.list_fields_export_capft_supports[col+14]] = sheet.cell(column=col, row=row).value


                            if sheet.cell(column=col, row=row).fill.fgColor.rgb == 'FFFF0000':
                                support['controle_calcul'] = 'nok'

                            else:
                                support['controle_calcul'] = 'ok'


                        if self.list_fields_export_capft_supports[col+14] not in ['Latitude (WGS84)', 'Longitude (WGS84)', 'Effort disponible après ajout câble']:
                            support[self.list_fields_export_capft_supports[col+14]] = str(sheet.cell(column=col, row=row).value).replace('\'', '').replace(u"\u2018", "'").replace(u"\u2019", "'").replace(u"\u0153", "oe")


                    if support[self.list_fields_export_capft_supports[16]] not in  ('EDF','ORT'):
                        array_result_bool = []


                        support['modele'] = support['Type d appui (format GESPOT)']
                        support['sup_remp'] = support['Type d appui aprés travaux']

                        if support['sup_remp'] == 'None' and support['controle_calcul'] == 'nok':
                            support['propriet'] = GESTIONNAIRE
                            support['com_etat'] = 'Implantation'


                        if support['Nature des travaux']:
                            support['com_etat'] = support['Nature des travaux']


                        for field in support:
                            if field in list_controle:
                                if support[field] == 'Oui':
                                    array_result_bool.append(True)
                                else:
                                    array_result_bool.append(False)

                        result = array_result_bool.count(array_result_bool[0]) == len(array_result_bool)
                        if result:
                            support['controle_visuel'] = 'ok'
                            support['commentair'] = 'Controle visuel OK'

                        if support['com_etat'] == 'None':
                            support['etat'] = 'ok'
                        else:
                            support['etat'] = 'nok'


                        if support['Forfait optique'] == 'Oui':
                            support['ban_vert'] = 'ok'
                        else:
                            support['ban_vert'] = 'nok'

                        if support['Installation réhausse'] == 'Oui':
                            support['h_traverse'] = 'Réhausse'

                        support['Longitude (WGS84)'] = support['Longitude (WGS84)'].replace('O','W')

                        if '°' in support['Latitude (WGS84)']:
                            if support['Latitude (WGS84)'][-1] != 'N':
                                support['Latitude (WGS84)'] = support['Latitude (WGS84)'] + 'N'
                        else:
                            neg, deg, mnt, sec = decdeg2dms(support['Latitude (WGS84)'].replace(',', '.'))
                            support['Latitude (WGS84)'] = str(abs(int(deg)))+ "°"+ str(abs(int(mnt)))+ "'"+ str(abs(float(sec[:6]))).replace('.',',')+ '"N'

                        if support['Longitude (WGS84)'][0] == '-' and  '°' in support['Longitude (WGS84)']:
                            if support['Longitude (WGS84)'][-1] != 'E':
                                support['Longitude (WGS84)'] = support['Longitude (WGS84)'] + 'E'
                            support['Longitude (WGS84)'] = support['Longitude (WGS84)'].replace("\'\'", '"')
                            print(re.split('[°\'"]', support['Longitude (WGS84)']))
                            deg, minutes, seconds, direction =  re.split('[°\'"]', support['Longitude (WGS84)'])
                            dd = (float(deg) + float(minutes)/60 + float(seconds.replace(',','.'))/(60*60)) * (-1 if direction in ['E', 'S'] else 1)
                            mnt,sec = divmod(dd*3600,60)
                            deg,mnt = divmod(mnt,60)
                            support['Longitude (WGS84)'] = str(abs(int(deg)))+ "°"+ str(abs(int(mnt)))+ "'"+ str(abs(float(sec))).replace('.',',')+ '"W'


                        else:
                            if '°' in support['Longitude (WGS84)']:
                                if support['Longitude (WGS84)'][-1] not in ('O','W','E'):
                                    support['Longitude (WGS84)'] = support['Longitude (WGS84)'] + 'E'
                            else:
                                neg, deg, mnt, sec = decdeg2dms(support['Longitude (WGS84)'].replace(',', '.'))
                                if neg:
                                    support['Longitude (WGS84)'] = str(abs(int(deg)))+ "°"+ str(abs(int(mnt)))+ "'"+ str(abs(float(sec[:6]))).replace('.',',')+ '"W'
                                else:
                                    support['Longitude (WGS84)'] = str(abs(int(deg)))+ "°"+ str(abs(int(mnt)))+ "'"+ str(abs(float(sec[:6]))).replace('.',',')+ '"E'


                        list_supports.append(support)


            self.list_etudes_capft.append({'Etude' : etude_name, 'Supports' : list_supports})



    class buildEtudeComacItem:

#         proj_wgs93 = pyproj.Proj(proj='latlong', ellps='epsg:2154')

        def __init__(self, pcm, list_etudes, list_fields_export_supports, type_supports,repare_pcm,option_boitier_fibre):


            self.type_supports = type_supports
            self.list_etudes_comac = list_etudes
            self.list_fields_export_supports = list_fields_export_supports
            self.pcm = pcm
            self.xml =  self.pcm['xml']
            self.directory = self.pcm['directory']
            self.name_1 = Path(self.pcm['directory'])
            self.name_1 = self.name_1.name
            self.name_2 = Path(self.pcm['path'])
            self.name_2 = self.name_2.name
            self.repare_pcm = repare_pcm
            self.option_boitier_fibre = option_boitier_fibre

            if re.search('(?<=-)([23D]*(-){1})', self.name_1) or re.search('((?<=-)(V){1}[0-9]{1})', self.name_1) or re.search('(?<=-)([A-CE-Z]{1})([0-9]{1,2})', self.name_1):
                self.name = self.name_1

            else:
                self.name = self.name_2.replace('.', '')


            self.list_item_etude = {}


            self.create_etude()








        def get_pdf(self):

            list_pdf = []
            for subdir, dirs, files in os.walk(self.directory):
                for file in files:

                    filepath = subdir + os.sep + file

                    if filepath.endswith(".pdf") and self.list_item_etude['Numero'] in filepath and self.pcm['pmz'] in filepath:
                        list_pdf.append({'path': filepath, 'modified':os.path.getmtime(filepath), 'size' : os.path.getsize(filepath)})


                list_pdf_sorted = sorted(list_pdf, key=lambda k: (k['size'], k['modified']), reverse=True)

            return (list_pdf_sorted[0]['path'].replace('\\\\?\\', '') if list_pdf_sorted else None)



        def create_etude(self):



                self.list_item_etude['id'] = self.pcm['id']
                self.list_item_etude['Etude'] = re.sub('([ a-zA-Z]*)$', '',re.sub('(_V){1}[0-9]{1}', '',re.sub('(-V){1}[0-9]{1}', '' , self.name)))
                num = re.search('(?<=-)([A-Z-]*)([0-9]{1,2})$', self.list_item_etude['Etude'])[0] if re.search('(?<=-)([A-Z-]*)([0-9]{1,2})$', self.list_item_etude['Etude']) else 'NONE'
                if re.search('((PA){1})', num) or re.search('((PB){1})', num) : num = 'NONE'
                self.list_item_etude['Etude'] = self.list_item_etude['Etude'].replace("\'", "")
                self.list_item_etude['Numero'] = re.sub('([_ -]*)', '', num) if num != 'NONE' else '-'
                self.list_item_etude['NRO'] = self.pcm['nro']
                self.list_item_etude['PMZ'] = self.pcm['pmz']
                self.list_item_etude['PA'] = re.search('(?<=PA)[0-9]*(?=-)', self.list_item_etude['Etude'])[0]  if re.search('(?<=PA)[0-9]*(?=-)', self.list_item_etude['Etude']) else None
                self.list_item_etude['PB'] = re.search('(?<=PB)[0-9]*(?=-)', self.list_item_etude['Etude'])[0]  if re.search('(?<=PB)[0-9]*(?=-)', self.list_item_etude['Etude']) else None
                self.list_item_etude['Version'] = self.pcm['version']
                self.list_item_etude['Type'] =  re.search('(?<=-)([23D]*)(?=-)', self.list_item_etude['Etude'])[0] if re.search('(?<=-)([23D]*)(?=-)', self.list_item_etude['Etude']) else None
                self.list_item_etude['Commune'] = self.xml.find("Commune").text if self.xml.find("Commune") is not None else ''
                self.list_item_etude['Insee'] = self.xml.find("Insee").text if self.xml.find("Insee") is not None else ''
                self.list_item_etude['Rue'] = self.xml.find("Rue").text if self.xml.find("Rue") is not None else ''
                self.list_item_etude['Portee'] = self.import_portees()
                self.list_item_etude['Lignes'] = self.import_lignes()
                self.list_item_etude['Supports'] = self.import_supports()
                self.list_item_etude['Infos'] = self.calcul_infos()
                self.list_item_etude['Path'] = self.directory.replace('\\\\?\\', '')
                self.list_item_etude['Pcm'] = self.pcm['path'].replace('\\\\?\\', '')
                self.list_item_etude['Pdf'] = self.get_pdf()
                print(self.list_item_etude['Etude'])

                if self.repare_pcm:

                    self.count = 0
                    self.count_to_repare = 0
                    self.count_ref = 0

                    for support in self.list_item_etude['Supports']:
                        if (support['X'] == '0' and support['Y'] == '0') or (support['X'] is None and support['Y'] is None):
                            self.count_to_repare += 1

                    for support in self.list_item_etude['Supports']:
                        if ((support['X'] != '0' and support['Y'] != '0') and (support['X'] is not None and support['Y'] is not None)) and (len(support['X']) > 7 or len(support['Y']) > 7):
                            self.count_ref += 1

                    print(self.count_to_repare)
                    if self.count != self.count_to_repare and self.count_ref != 0:
                        while self.count != self.count_to_repare:
                            self.repare_pcm_coord()

                        self.update_pcms_appuis()

                self.list_etudes_comac.append(self.list_item_etude)






        def update_pcms_appuis(self):
            tree = ET.parse(self.list_item_etude['Pcm'])
            root = tree.getroot()
            for support in self.list_item_etude['Supports']:
                for support_ in root.findall('Supports//Support'):
                    for field in support_:
                        if field.tag == 'Nom':
                            if field.text == support['Nom']:
                                for field_ in support_:
                                    if field_.tag == 'X':
                                        field_.text = support['X']
                                    elif field_.tag == 'Y':
                                        field_.text = support['Y']
            tree.write(self.list_item_etude['Pcm'])




        def repare_pcm_coord(self):



            for support_ in self.list_item_etude['Supports']:
                if (support_['X'] != '0' and support_['Y'] != '0') and (support_['X'] is not None and support_['Y'] is not None):

                    #print('ref',support_['Nom'])
                    for portee in self.list_item_etude['Portee']:
                        if portee['SuppG'] == support_['Nom'] or portee['SuppD'] == support_['Nom']:
                            for support in self.list_item_etude['Supports']:
                                if (support['X'] == '0' and support['Y'] == '0') or (support['X'] is None and support['Y'] is None):
                                    if portee['SuppD'] == support['Nom'] or portee['SuppG'] == support['Nom']:

                                        if float(portee['Angle'].replace(',', '.')) < 0:
                                            angle = 400 + float(portee['Angle'].replace(',', '.'))
                                        else:
                                            angle = float(portee['Angle'].replace(',', '.'))


                                        angle_hypo =  angle /400 * 360
                                        distance_y = float(portee['Longueur'].replace(',', '.')) * math.sin(math.radians(angle_hypo))
                                        distance_x = float(portee['Longueur'].replace(',', '.')) * math.cos(math.radians(angle_hypo))

                                        if (portee['SuppD'] == support['Nom']):
                                            support['X'] = float(support_['X'].replace(',', '.')) + distance_x
                                            support['Y'] = float(support_['Y'].replace(',', '.')) + distance_y
                                        if (portee['SuppG'] == support['Nom']):
                                            support['X'] = float(support_['X'].replace(',', '.')) - distance_x
                                            support['Y'] = float(support_['Y'].replace(',', '.')) - distance_y
                                        support['X'] = str(support['X']).replace('.', ',')
                                        support['Y'] = str(support['Y']).replace('.', ',')
                                        self.count  += 1
                                        print(support['Nom'],support['X'],support['Y'])







        def calcul_longueur_pose(self, portees):
            longueur = float(0)

            for key in portees:
                 longueur += float(portees[key][0])

            return longueur



        def calcul_longueur_totale_pose(self, lignes):
            longueur = float(0)
            for ligne in lignes:
                longueur += float(ligne['Longueur'])

            return longueur


        def calcul_longueur_totale_facture(self, fibre, list_supports):


            longueur = float(0)
            list_portees = []
            for k, ligne in enumerate(fibre['Lignes']):
                for portee in ligne['Portees']:

                    if k == 0:
                        list_portees.append(portee)
                        if len([(match) for match in list_supports['D2'] if match in portee]) == 1:
                            longueur += float(ligne['Portees'][portee][0])/2
                        if len([(match) for match in list_supports['D2'] if match in portee]) == 2:
                            longueur += float(ligne['Portees'][portee][0])

                    else:

                        portee_arr = portee.split(' - ')

                        i = len(list_portees)
                        ok = True
                        while i > 0:
                            if (list_portees[i-1].find(portee_arr[0]) >= 0) and (list_portees[i-1].find(portee_arr[1]) >= 0):
                                ok = False
                            i -= 1

                        if ok:
                            list_portees.append(portee)
                            if len([(match) for match in list_supports['D2'] if match in portee]) == 1:
                                longueur += float(ligne['Portees'][portee][0])/2
                            if len([(match) for match in list_supports['D2'] if match in portee]) == 2:
                                longueur += float(ligne['Portees'][portee][0])


            longueur_formule = '=INDIRECT(ADRESSE(LIGNE();COLONNE()-2;4))*40+' + str(longueur).replace('.', ',')
            longueur_abs = (len(list_supports['D3']) + len(list_supports['D3-1'])) * 40 + longueur

            return {'Absolue' : str(longueur_abs).replace('.', ','), 'Formule' : longueur_formule}









        def geobuffer_athd(self, athd, list_portee_bt, list_supports_bt, supports_bt, supports_bt_nok):






            athd_portee = {}
            adjacents = []
            adjacents_portees = {}
            matchers_nature = ['FF', 'FM','FT']
            athd_modele = ''

            nom_comac = str(athd['Hauteur']) + str(athd['Nature']) + str(athd['Classe'])

            for type_support in self.type_supports:
                if nom_comac == (str(type_support['Hauteur totale'])+str(type_support['Nature'])+str(type_support['Classe'])):
                    if len(type_support['NomGESPOT'].split(';')) > 1:
                        athd_modele = type_support['NomGESPOT'].split(';')[0]
                    else:
                        athd_modele = type_support['NomGESPOT']
                    break



            lat = float(athd['Y'].replace(',','.'))
            lng = float(athd['X'].replace(',','.'))


            athd_point_buffer = geometry.Point(lng,lat ).buffer(10)
            for support in supports_bt_nok:
                if athd_point_buffer.contains(geometry.Point(float(support['X'].replace(',','.')), float(support['Y'].replace(',','.')))):


                    self.update_troncon_nok(support['Nom'])
                    return 'Renfort', support['numero'], athd_modele





            for ligne in self.list_item_etude['Lignes']['Fibre']['Lignes']:
                for portee, values in ligne['Portees'].items():

                    if athd['Nom'] in portee:
                        for supp in supports_bt:
                            if supp['Nom'] in portee and supp['Nom'] != athd['Nom']:
                                adjacents.append(supp['Nom'])
                                athd_portee[supp['Nom']] = { portee : ligne['Portees'][portee]}




            for adj in adjacents:
                adjacents_portees[str(adj)] = []
                for ligne in self.list_item_etude['Lignes']['BT']['Lignes']:
                    for portee, values in ligne['Portees'].items():
                        if adj in portee:
                            adjacents_portees[str(adj)].append({ portee : ligne['Portees'][portee]})



            for key_athd, values_athd in athd_portee.items():
                for portee, values in athd_portee[str(key_athd)].items():
                    athd_distance = athd_portee[str(key_athd)][str(portee)][0]
                    if portee.find(athd['Nom']) == 0:
                        athd_angle = athd_portee[str(key_athd)][str(portee)][1]
                    else:
                        athd_angle = athd_portee[str(key_athd)][str(portee)][1] + 200


                for portees_bt in adjacents_portees[str(key_athd)]:
                    for portee, values in portees_bt.items():
                        adj_distance = portees_bt[str(portee)][0]
                        if portee.find(key_athd) != 0:
                            adj_angle = portees_bt[str(portee)][1]
                        else:
                            adj_angle = portees_bt[str(portee)][1] + 200



                        remplacement = portee.replace(str(key_athd), '')
                        remplacement = remplacement.replace(' - ', '')

                        remplacement = re.sub('(NC){0,1}(IN){0,1}(D[0-9]{1}){0,1}(FT[0-9]-{0,1}){0,1}([ -]*)', '', remplacement)
                        remplacement =  re.sub('[0-9]*E', 'E', remplacement).strip()


                        if (abs(athd_distance - adj_distance) < 15 and abs(athd_angle - adj_angle) < 10):

                            self.update_troncon_nok(remplacement)
                            return 'Renfort', remplacement, athd_modele





            if 'Raison_ATHD' not in self.list_fields_export_supports:
                self.list_fields_export_supports.append('Raison_ATHD')
            return 'Portée', 'NONE', athd_modele


        def update_troncon_nok(self, supp_nom):

            for support in self.list_item_etude['Supports']:
                if support['Nom'] == supp_nom:
                    support['troncon'] = 'D2'


        def calcul_appuis_calcules(self):
            liste_appuis_supports_D2 = []
            liste_appuis_supports_D3 = []
            liste_appuis_supports_D3_1 = []
            liste_appuis_supports_ATHD_renfort = []
            liste_appuis_supports_ATHD_portee = []
            liste_appuis_supports_D2_capft_remplace = []
            liste_appuis = {}
            list_portee_bt = []
            list_portee_bt_capft = []
            list_supports_bt = []
            supports_bt = []
            supports_bt_nok = []
            supports = []
            # list_update_supports = {}
            if 'Raison_ATHD' not in self.list_fields_export_supports:
                self.list_fields_export_supports.append('Raison_ATHD')
            if 'Remplace' not in self.list_fields_export_supports:
                self.list_fields_export_supports.append('Remplace')

            for ligne in self.list_item_etude['Lignes']['BT']['Lignes']:
                for Key, portee in ligne['Portees'].items():
                    list_portee_bt.append(Key)


            list_supports_bt = self.list_item_etude['Lignes']['BT']['Supports_list']
            for support in self.list_item_etude['Supports']:
                if support['Nom'] in list_supports_bt and support['NonCalcule'] == '0':
                    supports_bt.append(support)
                if support['Nom'] in list_supports_bt and support['NonCalcule'] == '1':
                    supports_bt_nok.append(support)



            for support in self.list_item_etude['Supports']:

                if support['troncon'] == 'D2' and support['etat'] == 'ok' and support['propriet'] == 'ENEDIS':
                    liste_appuis_supports_D2.append(support['Nom'])
                    liste_appuis_supports_D2_capft_remplace.append(support['Nom'])

                if support['troncon'] == 'D3' and support['etat'] == 'ok' and support['propriet'] == 'ENEDIS':
                    liste_appuis_supports_D3.append(support['Nom'])

                if support['troncon'] == 'D2' and support['propriet'] == GESTIONNAIRE:
                    raison, remplacement, athd_modele = self.geobuffer_athd(support, list_portee_bt, list_supports_bt, supports_bt, supports_bt_nok)
                    if raison == 'Renfort':
                        liste_appuis_supports_ATHD_renfort.append(support['Nom'])
                        liste_appuis_supports_D2_capft_remplace.append(remplacement)
                    if raison == 'Portée':
                        liste_appuis_supports_ATHD_portee.append(support['Nom'])

                    # try:
                    support['ban_vert'] = None
                    support['Raison_ATHD'] = raison
                    insee = self.list_item_etude['Insee']
                    if insee is None:
                        insee = ''
                    support['Remplace'] = (str(remplacement) + '/' + insee) if remplacement != 'NONE' else None
                    support['modele'] = athd_modele
                    if raison == 'Renfort':
                        support['numero'] = support['numero'] + '|' + support['Remplace']
                    if 'modele' not in self.list_fields_export_supports:
                            self.list_fields_export_supports.append('modele')
                        # list_update_supports[support['Remplace']] = {'Remplace': support['Nom'], 'modele': athd_modele}

                    # except:
                    #     print(self.list_item_etude['Etude'])





            supports = self.list_item_etude['Supports']

            liste_appuis_supports_fibre = self.list_item_etude['Lignes']['Fibre']['Supports_list']
            liste_appuis_supports_telecom = self.list_item_etude['Lignes']['Telecom']['Supports_list']
            for supp in liste_appuis_supports_telecom:
                if supp in liste_appuis_supports_fibre:
                    for support in supports:
                        if supp == support['Nom']:

                            lat = float(support['Y'].replace(',','.'))
                            lng = float(support['X'].replace(',','.'))

                            ft_point_buffer = geometry.Point(lng,lat ).buffer(5)
                            for bt_nok in supports_bt_nok:
                                if ft_point_buffer.contains(geometry.Point(float(bt_nok['X'].replace(',','.')), float(bt_nok['Y'].replace(',','.')))):
                                    liste_appuis_supports_D2_capft_remplace.append(bt_nok['Nom'])

                    liste_appuis_supports_D2_capft_remplace.append(supp)


            for ligne in self.list_item_etude['Lignes']['BT']['Lignes']:
                for portee in ligne['Portees']:
                    portee_arr = portee.split(' - ')
                    for supp in portee_arr:
                        if supp in liste_appuis_supports_D2_capft_remplace and portee not in list_portee_bt_capft:
                            list_portee_bt_capft.append(portee)

            for ligne in self.list_item_etude['Lignes']['Telecom']['Lignes']:
                for portee in ligne['Portees']:
                    portee_arr = portee.split(' - ')
                    for supp in portee_arr:
                        if supp in liste_appuis_supports_D2_capft_remplace and portee not in list_portee_bt_capft:
                            list_portee_bt_capft.append(portee)





            for support_d3 in liste_appuis_supports_D3:

                for portee in list_portee_bt_capft:
                    portee_arr = portee.split(' - ')

                    if support_d3 in portee_arr:
                        for support in liste_appuis_supports_D2_capft_remplace:
                            if support in portee_arr:

                                for supp in self.list_item_etude['Supports']:
                                    if supp['Nom'] == support_d3 and support_d3 not in liste_appuis_supports_D3_1:
                                        supp['troncon'] = 'D3-1'
                                        liste_appuis_supports_D3_1.append(support_d3)

            try:
                for supp in liste_appuis_supports_D3_1:
                    liste_appuis_supports_D3.remove(supp)

            except:
                pass





            liste_appuis["D2"] = liste_appuis_supports_D2
            liste_appuis["D3"] = liste_appuis_supports_D3
            liste_appuis["D3-1"] = liste_appuis_supports_D3_1
            liste_appuis["ATHD_Renfort"] = liste_appuis_supports_ATHD_renfort
            liste_appuis["ATHD_Portée"] = liste_appuis_supports_ATHD_portee

            return liste_appuis



        def calcul_appuis_classification(self, support):



            list_supports_total_fibre = self.list_item_etude['Lignes']['Fibre']['Supports_list']

            troncon = None
            matchers_name = ['NC','IN','FAC','POT']
            matchers_nature = ['FF', 'FM','FT']
            erreur = ''
            if support['Nature'] is None: support['Nature'] = ''
            if support['Nom']  is None: support['Nom'] = ''

            troncon = ''
            if [(match) for match in matchers_nature if match in support['Nature']]:
                support['propriet'] = 'ORANGE'
            else:
                support['propriet'] = 'ENEDIS'


            if re.search('E[0-9]*', support['Nom']): troncon = 'D3'

            if support['NonCalcule'] == '0' and not [(match) for match in matchers_nature if match in support['Nature']] and (support['Nom'] in list_supports_total_fibre) and not [(match) for match in matchers_name if match in support['Nom']]:
                troncon = 'D2'
                support['etat'] = 'ok'
                if support['ban_vert'] == 'nok': erreur = 'Bandeau vert non coché'


            if not [(match) for match in matchers_nature if match in support['Nature']] and (support['Nom'] not in list_supports_total_fibre) and not [(match) for match in matchers_name if match in support['Nom']]:
                if ((support['optBandeauVertAPoser'] == '1' if 'optBandeauVertAPoser' in support else False) or (support['optBandeauVertExistant'] == '1' if 'optBandeauVertExistant' in support else False)) and (support['optBoitierFibre'] == '0' if 'optBoitierFibre' in support else False):
                    troncon = 'D3'
                    support['etat'] = 'ok'
                    if support['ban_vert'] == 'nok': erreur = 'Bandeau vert non coché'
                if support['optBoitierFibre'] == '1':
                    troncon = 'D2'
                    support['etat'] = 'ok'
                    if support['ban_vert'] == 'nok': erreur = 'Bandeau vert non coché'

            if [(match) for match in matchers_nature if match in support['Nature']] and (support['APoser'] == '1' or 'THD' in support['Nom']) and support['Nom'] in list_supports_total_fibre:
                support['APoser'] =  1
                support['Raison_ATHD'] = ''
                support['etat'] = 'ok'
                support['propriet'] = GESTIONNAIRE
                support['commentair'] = support['Nom']
                troncon = 'D2'


            if support['propriet'] == 'ENEDIS':
                len_num = len(support['numero'])
                if support['numero'][:1] == 'E' and len_num < 7 and len_num > 5:
                    support['numero'] = support['numero'].replace('E', '')

                    while len_num < 7:
                        support['numero'] = '9' + support['numero']
                        len_num += 1
                    support['numero'] = 'E' + support['numero']

            if  support['optBoitierFibre'] == '1' and support['Nom'] not in list_supports_total_fibre:
                
                if self.option_boitier_fibre:
                    troncon = 'D3'
                else:
                    troncon = 'D2'



            return erreur, troncon




        def calcul_infos(self):
            list_infos = {}

            list_infos['Listes_appuis_calculés'] = self.calcul_appuis_calcules()

            list_infos['Nombre_D2'] = len(list_infos['Listes_appuis_calculés']['D2'])
            list_infos['Nombre_D3'] = len(list_infos['Listes_appuis_calculés']['D3'])+len(list_infos['Listes_appuis_calculés']['D3-1'])
            list_infos['Nombre_ATHD_Portée'] = len(list_infos['Listes_appuis_calculés']['ATHD_Portée'])
            list_infos['Nombre_ATHD_Renfort'] = len(list_infos['Listes_appuis_calculés']['ATHD_Renfort'])

            list_infos['Longueur_totale_pose'] = str(self.list_item_etude['Lignes']['Fibre']['Longueur_totale']).replace('.', ',')
            list_infos['Longueur_totale_facture'] = self.calcul_longueur_totale_facture(self.list_item_etude['Lignes']['Fibre'], list_infos['Listes_appuis_calculés'])


            count_appuis = [list_infos['Nombre_D2'], list_infos['Nombre_D3'], list_infos['Nombre_ATHD_Portée'], list_infos['Nombre_ATHD_Renfort']]
            if (count_appuis[0] > 0 and count_appuis[1] < 1) or (count_appuis[2] > 0 or count_appuis[3] > 0):
                list_infos['Type_calculé'] = 'D2'

            if count_appuis[0] < 1 and count_appuis[1] > 0:
                list_infos['Type_calculé'] = 'D3'

            if count_appuis[0] > 0 and count_appuis[1] > 0:
                list_infos['Type_calculé'] = 'D2D3'

            if count_appuis[0] < 1 and count_appuis[1] < 1 and count_appuis[2] < 1 and count_appuis[3] < 1:
                list_infos['Type_calculé'] = 'Annulé'



            return list_infos



        def import_supports(self):

            def search(values, searchFor):
                match = []
                for k in values:
                    if searchFor in k:
                            match.append(k)
                return match


            list_supports =[]

            for support in self.xml.findall('Supports//Support'):

                list_item_support = {}

                list_item_support['NonCalcule'] = ''
                list_item_support['optBoitierFibre'] = ''
                list_item_support['commentair'] = ''




                for field in support:

                    list_item_support[field.tag] = field.text
                    if field.tag not in list_item_support:
                        list_fields_support.append(field.tag)

                    if field.tag not in self.list_fields_export_supports:
                        self.list_fields_export_supports.append(field.tag)


                if list_item_support['Nom'] is not None:
                    list_item_support['numero'] = re.sub('(NC){0,1}(IN){0,1}(D[0-9]{1}){0,1}(FT[0-9]-{0,1}){0,1}([ -]*)', '', list_item_support['Nom'])
                    list_item_support['numero'] =  re.sub('[0-9]*E', 'E', list_item_support['numero']).strip()
                else:
                    list_item_support['numero'] = ''





                iterate_fields = []
                for match in search(list_item_support, 'TraverseAPoser'):
                    iterate_fields.append(match)


                if len(iterate_fields) > 0:
                    list_item_support['h_traverse'] = list_item_support[iterate_fields[-1]]
                    if 'h_traverse' not in self.list_fields_export_supports:
                        self.list_fields_export_supports.append('h_traverse')

                iterate_fields = []

                list_item_support['ban_vert'] = 'nok'
                for match in search(list_item_support, 'optBandeauVert'):
                    if list_item_support[match] == '1':
                        list_item_support['ban_vert'] = 'ok'
                        break

                list_item_support['etat'] = 'nok'




                if 'ban_vert' not in self.list_fields_export_supports:
                        self.list_fields_export_supports.append('ban_vert')
                if 'etat' not in self.list_fields_export_supports:
                        self.list_fields_export_supports.append('etat')
                if 'propriet' not in self.list_fields_export_supports:
                        self.list_fields_export_supports.append('propriet')
                if 'Erreur' not in self.list_fields_export_supports:
                        self.list_fields_export_supports.append('Erreur')



                list_item_support['Erreur'], list_item_support['troncon'] = self.calcul_appuis_classification(list_item_support)
                list_supports.append(list_item_support)

            sorted_list_supports = []

            sorted_list_supports = sorted(list_supports, reverse=True, key=lambda k: (k['etat'],k['troncon']))

            return sorted_list_supports




        def import_portees(self):

            list_portees = []

            for support in self.xml.findall('Portees//Portee'):

                list_item_portee = {}


                for field in support:

                    list_item_portee[field.tag] = field.text
                    if field.tag not in list_item_portee:
                        list_item_portee.append(field.tag)

                list_portees.append(list_item_portee)


            return list_portees






        def build_ligne(self, ligne):
            list_fields_lignes = []
            list_item = {}
            list_supports_ligne = []
            list_supports = []
            list_couples = []
            list_portee = {}


            for field in ligne:

                if field.tag == 'Supports':

                    temp = []

                    len_supp = len(field.findall('Support'))
                    for supp in field.iter('Support'):
                        list_supports.append(supp.text)
                        temp.append(supp.text)

                        if len(temp) % 3 == 0:
                            temp.pop(0)

                        if len(temp) % 2 == 0:
                            list_couples.append(' - '.join(temp))




                list_item['Supports'] = list_supports

                if field.tag == 'Portees':
                    for couple in list_couples:
                        for i, portee in enumerate(self.list_item_etude['Portee']):
                            if portee['SuppG'] in couple and portee['SuppD'] in couple:
                                if couple.find(portee['SuppG']) == 0:
                                    list_portee[couple] = [float(portee['Longueur'].replace(',', '.')), float(portee['Angle'].replace(',', '.'))]
                                else:
                                    list_portee[couple] = [float(portee['Longueur'].replace(',', '.')), float(portee['Angle'].replace(',', '.')) + 200]





                    list_item['Portees'] = list_portee

                else:
                    list_item[field.tag] = field.text

                if field.tag not in list_fields_lignes and field.tag != 'Supports':
                    list_fields_lignes.append(field.tag)

            list_item['Longueur'] = self.calcul_longueur_pose(list_portee)

            return list_item








        def import_lignes(self):

            list_lignes_fibre = []
            list_lignes_telecom = []
            list_lignes_bt = []
            list_supports_total_fibre = []
            list_supports_total_telecom = []
            list_supports_total_bt = []


            for ligne in self.xml.iter('LigneTCF'):



                if ligne.find('Cable').text.find('L1') > -1  and ligne.find('APoser').text == '1':
                    ligne_ = self.build_ligne(ligne)

                    list_lignes_fibre.append(ligne_)
                    for support in ligne_['Supports']:
                        if support not in list_supports_total_fibre:
                            list_supports_total_fibre.append(support)


                else:
                    ligne_ = self.build_ligne(ligne)

                    list_lignes_telecom.append(ligne_)
                    for support in ligne_['Supports']:
                        if support not in list_supports_total_telecom:
                            list_supports_total_telecom.append(support)




            for ligne in self.xml.iter('LigneBT'):

                ligne_ = self.build_ligne(ligne)

                list_lignes_bt.append(ligne_)
                for support in ligne_['Supports']:
                    if support not in list_supports_total_bt:
                        list_supports_total_bt.append(support)


            return {'Fibre': {'Supports_list': list_supports_total_fibre, 'Longueur_totale' : self.calcul_longueur_totale_pose(list_lignes_fibre), 'Lignes' : list_lignes_fibre},
                    'Telecom': {'Supports_list': list_supports_total_telecom, 'Longueur_totale' : self.calcul_longueur_totale_pose(list_lignes_telecom), 'Lignes' : list_lignes_telecom},
                    'BT': {'Supports_list': list_supports_total_bt, 'Longueur_totale' : self.calcul_longueur_totale_pose(list_lignes_bt), 'Lignes' : list_lignes_bt}}
