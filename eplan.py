import unittest
import csv
import openpyxl
import os
import re
from os.path import basename
from zipfile import ZipFile
from PIL import Image
from . import utils as utils
from .utils import *
import sys
import time
import numpy as np
import scipy.interpolate as si
from . import pdf_compressor
from . import create_doc
from datetime import datetime
from time import sleep, time
from random import uniform, randint
import threading
from qgis.PyQt.QtCore import *
from qgis.PyQt.QtGui import *
from qgis.PyQt.QtWidgets import *
from qgis.core import *
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.proxy import Proxy, ProxyType
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
#from selenium_stealth import stealth
import urllib.request

class ePlanDriver:

    def __init__(self, email, password):

        self.email = email
        self.password = password
        self.headless = False
        self.options = None
        self.profile = None
        self.capabilities = None
        self.driver = None
        self.data = []
        self.get_data()
        self.setUp()





    def setUpOptions(self):
        self.options = webdriver.ChromeOptions()
        self.options.add_argument("start-maximized")
        self.options.add_experimental_option("excludeSwitches", ["enable-automation"])
        self.options.add_experimental_option('useAutomationExtension', False)
        self.options.add_argument('--disable-blink-features=AutomationControlled')
        self.options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36')

        prefs = {"download.default_directory" : utils.DIR_OUTPUT_}
        self.options.add_experimental_option("prefs",prefs)
        self.options.headless = self.headless

    # Setup profile with buster captcha solver
    def setUpProfile(self):
        #self.profile = webdriver.ChromeProfile()
        #self.profile.set_preference('xpinstall.signatures.required', False)
        #self.profile._install_extension("buster_captcha_solver_for_humans-1.1.0-an+fx.xpi", unpack=False)
        #self.profile.set_preference("security.fileuri.strict_origin_policy", False)
        self.profile.update_preferences()


    # Enable Marionette, An automation driver for Mozilla's Gecko engine
    def setUpCapabilities(self):

        self.capabilities['marionette'] = True

    # Setup proxy
    def setUpProxy(self):
        self.log(PROXY)
        self.capabilities['proxy'] = {
            "proxyType": "MANUAL",
            "httpProxy": PROXY,
            "ftpProxy": PROXY,
            "sslProxy": PROXY
        }

    # Setup settings
    def setUp(self):
        #self.setUpProfile()
        self.setUpOptions()
        #self.setUpCapabilities()
        #self.setUpProxy() # comment this line for ignore proxy

        # On Linux?
        # https://github.com/mozilla/geckodriver/issues/1756
        # binary = FirefoxBinary('/usr/lib/firefox-esr/firefox-esr')
        # self.driver = webdriver.Firefox(options=self.options, capabilities=self.capabilities, firefox_profile=self.profile, executable_path='./geckodriver_linux', firefox_binary=binary)
        self.driver = webdriver.Chrome(options=self.options,executable_path=utils.DIR_PLUGIN + '\\' + 'chromedriver.exe')
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")


    def getDriver(self):
        return self.driver

    def getData(self):
        return self.data

    def authentificate(self):

        driver = self.driver

        self.log("Connexion à E-Plan")
        actions = ActionChains(driver)


        driver.get("https://www.e-plans.fr/")
        WebDriverWait(driver, timeout=3).until(EC.presence_of_element_located((By.XPATH,"/html/body/div[1]/div/div[2]/div/form/input[4]")))

        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div/form/input[3]").send_keys(self.email)
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div/form/input[4]").click()

        WebDriverWait(driver, timeout=6000).until(EC.presence_of_element_located((By.XPATH ,"/html/body/div[3]/div/div[2]/div/form/div[1]/div/input")))

        driver.find_element_by_xpath("/html/body/div[3]/div/div[2]/div/form/div[1]/div/input").send_keys(self.email)
        #WebDriverWait(driver, timeout=6000).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@src, 'https://www.google.com/recaptcha/api2/anchor')]")))
        #sleep(2)
        #checkbox = driver.find_element_by_xpath("//span[@id='recaptcha-anchor']")
        #label = driver.find_element_by_xpath('//*[@id="rc-anchor-container"]/div[3]/div[2]')

        #action = ActionChains(driver)
        #action.move_to_element(label)
        #action.pause(0.5)
        #action.move_to_element(checkbox)
        #action.pause(1)
        #action.click(checkbox)
        #action.pause(2)
        #action.perform()
        #driver.switch_to.default_content()
        WebDriverWait(driver, timeout=6000).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="idToken3_0"]'))).click()

        WebDriverWait(driver, timeout=6000).until(EC.visibility_of_element_located((By.XPATH,"/html/body/div[3]/div/div[2]/div/form/div[2]/div[1]/input")))
        driver.find_element_by_xpath("/html/body/div[3]/div/div[2]/div/form/div[2]/div[1]/input").send_keys(self.password)

        driver.find_element_by_xpath("/html/body/div[3]/div/div[2]/div/form/div[3]/div/input").click()
        self.log("Connecté")
        sleep(2)

    def log(s,t=None):
            now = datetime.now()
            if t == None :
                    t = "Main"
            print ("%s :: %s -> %s " % (str(now), t, s))


    def get_data(self):
        with open(utils.DIR_OUTPUT_ + '\\' + 'export_tableau_suivi_enedis.csv', newline='', encoding='windows-1252') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                self.data.append(row)

limit_size = 10000000


class ePlanAutoDeposeEtude(threading.Thread):

    def __init__(self,description,email, password,convention,type_etude, depose, excel):


        threading.Thread.__init__(self)

        self.convention = convention
        self.type_etude = type_etude
        self.depose = depose
        self.number = 0
        self.excel = excel
        self.ePlanDriver = ePlanDriver(email,password)
        self.driver = self.ePlanDriver.getDriver()
        self.data = self.ePlanDriver.getData()
        self.exception = None
        self.messages = []
        self.Terminated = False



    def run(self):


        self.ePlanDriver.authentificate()
        '''stealth(
            driver = self.driver,
            user_agent= "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36",
            languages = ["fr-FR", "fr"],
            vendor = "Google Inc.",
            platform = "Win32",
            webgl_vendor = "Intel Inc.",
            renderer = "Intel Iris OpenGL Engine",
            fix_hairline = True,
            run_on_insecure_origins = False
        )'''
        self.maj_list_convention()
        self.iterate_over_etudes()
        self.driver.quit()
        self.Terminated = True
        return True


    def log(s,t=None):
            now = datetime.now()
            if t == None :
                    t = "Main"
            print ("%s :: %s -> %s " % (str(now), t, s))

    def maj_list_convention(self):


        driver = self.driver
        driver.get("https://www.e-plans.fr/DemandeAppuisCommuns/Create")

        WebDriverWait(driver, timeout=30).until(EC.visibility_of_element_located((By.XPATH,"/html/body/div[1]/div[2]/div/form/div[2]/div/select")))

        select_element_convention = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/form/div[2]/div/select')

        options_dict = {}
        options = [x for x in select_element_convention.find_elements_by_tag_name("option")]
        for element in options:
            options_dict[element.text] = element.get_attribute("value")

        with open(DIR_PLUGIN + '/config/eplan_conventions.json',"w") as outfile:
            json.dump(options_dict, outfile)
        self.log("La liste des conventions a été mise à jour")


    def depose_etude(self,code_insee,num_affaire_operateur,adresse,longueur_facture,nb_d2,nb_d3,archive):

        driver = self.driver
        driver.get("https://www.e-plans.fr/DemandeAppuisCommuns/Create")

        WebDriverWait(driver, timeout=30).until(EC.visibility_of_element_located((By.XPATH,"/html/body/div[1]/div[2]/div/form/div[2]/div/select")))

        select_element_convention = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/form/div[2]/div/select')

        select_object_convention = Select(select_element_convention)
        select_object_convention.select_by_value(self.convention)

        WebDriverWait(driver, timeout=30).until(EC.visibility_of_element_located((By.XPATH,"/html/body/div[1]/div[2]/div/form/div[4]/div/select/option[1]")))


        select_element_type = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/form/div[3]/div/select')
        select_object_type = Select(select_element_type)
        select_object_type.select_by_visible_text(self.type_etude)

        select_element_insee = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/form/div[4]/div/select')
        select_object_insee = Select(select_element_insee)
        select_object_insee.select_by_value(code_insee)

        driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/form/div[5]/div/input").send_keys(num_affaire_operateur)
        driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/form/div[6]/div/input").send_keys(adresse)
        driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/form/div[7]/div/label[3]/input').click()

        WebDriverWait(driver, timeout=30).until(EC.visibility_of_element_located((By.XPATH,"/html/body/div[1]/div[2]/div/form/div[10]/div/input")))

        driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/form/div[10]/div/input").send_keys(longueur_facture)
        driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/form/div[11]/div/input").send_keys(nb_d2)
        driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/form/div[14]/div/input").send_keys(nb_d3)


        driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/form/div[8]/div/div[2]/div/input").send_keys(archive)
        WebDriverWait(driver, timeout=30).until(EC.visibility_of_element_located((By.XPATH,"/html/body/div[1]/div[2]/div/form/div[8]/div/div[3]/div/ul")))
        if self.depose == 2:
            driver.find_element_by_xpath('//*[@id="save"]').click()
            self.log("Etude " + num_affaire_operateur + " déposée")




    def resize_image(self,infile):
        size = 1200, 1200


        outfile = infile[:-4] +'_compressed'
        try:
            im = Image.open(infile)
            im.thumbnail(size, Image.ANTIALIAS)
            im.save(outfile, "JPEG")
            return outfile
        except IOError:
            print ("cannot create thumbnail for '%s'" % infile)


    def create_zip(self,row,archive):
        with ZipFile(archive, 'w') as zipObj:
        # Add multiple files to the zip
            size_limit_photo = 500000
            size_limit_pdf = 5000000
            root = row['path']

            dirPhotos = root + '\\' + row["Nom d'affaire E-PLAN"] + '_Photos'
            zipObj.write(row['pcm'],basename(row['pcm']))
            try:
                zipObj.write(dirPhotos,basename(dirPhotos))
                for folderName, subfolders, filenames in os.walk(dirPhotos):
                    for filename in filenames:
                        filePath = os.path.join(folderName, filename)
                        outfile = filePath
                        if os.stat(filePath).st_size > size_limit_photo:
                            print(filePath,os.stat(filePath).st_size)
                            outfile = self.resize_image(filePath)

                        zipObj.write(outfile,basename(dirPhotos) + '\\' + basename(filePath))
                        if outfile != filePath:
                            os.remove(outfile)
            except:
                print("Le dossier photos est manquant ou introuvable")
                return False

            for folderName, subfolders, filenames in os.walk(root):
                for filename in filenames:
                    filePath = os.path.join(folderName, filename)
                    if 'pdf' in filePath:
                        pdf_compressor.compress(filePath, filePath[:-4] + '_compressed.pdf', 3)

                        zipObj.write( filePath[:-4] + '_compressed.pdf',basename(filePath))
                        os.remove(filePath[:-4] + '_compressed.pdf')
            if self.excel == 2:
                for folderName, subfolders, filenames in os.walk(root):
                    for filename in filenames:
                        filePath = os.path.join(folderName, filename)
                        if 'xls' in filePath and 'ExportComac' in filePath:
                            zipObj.write(filePath,basename(filePath))

#        if os.stat(archive).st_size > limit_size:
#            command = 'gzip -c {} | split -b {} - compressed.gz'.format(archive, limit_size)
#            os.system(command)





    def iterate_over_etudes(self):




        for i, row in enumerate(self.data):
            if float(row['Longueur à facturer ENEDIS (absolu)'].replace(',','.')) == 0 or len(row['INSEE']) != 5:
                print(row["Nom d'affaire E-PLAN"], 'Etude annulée')
            else:
                while True:
                    try:

                        #if float(row['Longueur à facturer ENEDIS (absolu)']) > 0 and (row['INSEE'] is not None or row['INSEE'] != ''):
                        if not os.path.isdir(row['path'].replace(row["Nom d'affaire E-PLAN"], 'ARCHIVES_E_PLAN')):
                            os.mkdir(row['path'].replace(row["Nom d'affaire E-PLAN"], 'ARCHIVES_E_PLAN'))
                        archive = row['path'].replace(row["Nom d'affaire E-PLAN"], 'ARCHIVES_E_PLAN') + '\\' +  row["Nom d'affaire E-PLAN"] + '.zip'
                        #shutil.make_archive(archive, 'zip', row['path'])
                        archived = self.create_zip(row,archive)
                        if archived:
                            self.log("Archive " + archive + " créée")

                            if os.stat(archive).st_size < limit_size:
                                self.depose_etude(row['INSEE'],row["Nom d'affaire E-PLAN"],row['Adresse'],row['Longueur à facturer ENEDIS (absolu)'],row['Nb de Support Enedis'],row['Nb de Support D3'],archive)
                                #os.remove(archive)
                            else:
                                print(row["Nom d'affaire E-PLAN"], os.stat(archive).st_size)
                        else:
                            break    
                    except:
                        continue

                    self.log(str(i+1))
                    self.number += 1
                    break
        self.log(str(self.number) + " études ont été importées")







class ePlanAutoExportBrut(threading.Thread):

    def __init__(self,description,email, password, dr, operateur, departement, date):

        threading.Thread.__init__(self)

        self.ePlanDriver = ePlanDriver(email,password)
        #self.driver = self.ePlanDriver.getDriver()
        self.data = self.ePlanDriver.getData()
        self.dr = dr
        self.operateur = operateur
        self.departement = departement
        self.date = date
        self.exception = None
        self.messages = []
        self.Terminated = False



    def run(self):


        #self.ePlanDriver.authentificate()
        '''stealth(
            driver = self.driver,
            user_agent= "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36",
            languages = ["fr-FR", "fr"],
            vendor = "Google Inc.",
            platform = "Win32",
            webgl_vendor = "Intel Inc.",
            renderer = "Intel Iris OpenGL Engine",
            fix_hairline = True,
            run_on_insecure_origins = False
        )'''
        #self.maj_list_dr()
        #self.get_export()
        #self.link_export()
        self.iterate_over_etudes()
        #self.driver.quit()
        self.Terminated = True
        return True


    def log(s,t=None):
            now = datetime.now()
            if t == None :
                    t = "Main"
            print ("%s :: %s -> %s " % (str(now), t, s))

    def maj_list_dr(self):


        driver = self.driver
        driver.get("https://www.e-plans.fr/ExportAppuisCommun?typeExportAppuisCommuns=ExportBrut")

        select_element_dr = WebDriverWait(driver, timeout=30).until(EC.visibility_of_element_located((By.XPATH,"/html/body/div[1]/div[2]/div/form/div/div/div[1]/div[1]/div[2]/div/span/span[1]/span/ul")))



        options_dict = {}
        options = [x for x in select_element_dr.find_elements_by_tag_name("li")]
        for element in options:
            options_dict[element.get_attribute('title')] = element.get_attribute('title')

        with open(DIR_PLUGIN + '/config/eplan_dr.json',"w") as outfile:
            json.dump(options_dict, outfile)
        self.log("La liste des conventions a été mise à jour")



    def get_export(self):

        for file in os.listdir(utils.DIR_OUTPUT_):
            if 'ExportAppuisCommunBrut' in file:
                os.remove(utils.DIR_OUTPUT_ + file)

        driver = self.driver
        driver.get("https://www.e-plans.fr/ExportAppuisCommun?typeExportAppuisCommuns=ExportBrut")

        select_element_dr = WebDriverWait(driver, timeout=30).until(EC.visibility_of_element_located((By.XPATH,"/html/body/div[1]/div[2]/div/form/div/div/div[1]/div[1]/div[2]/div/span/span[1]/span/ul")))

        try:
            while select_element_dr.find_element_by_xpath("//li[@class='select2-selection__choice']") is not None:
                try:

                    span = select_element_dr.find_element_by_xpath("//li[@class='select2-selection__choice']").find_element_by_xpath("//span[@class='select2-selection__choice__remove']")

                    span.click()



                except:
                    continue
        except:
            print('finish')


        select_element_dr = driver.find_element_by_xpath('//select[@id="Filtreur_FiltrePlaqueErdf"]')
        select_object_dr = Select(select_element_dr)
        select_object_dr.select_by_visible_text(self.dr)


        driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/form/div/div/div[2]/input').click()
        sleep(2)
        driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/p/a').click()
        sleep(2)

    def link_export(self):

        list_etudes_export_brut = []
        dict_etude_export_brut = {}
        list_field_etude_export_brut = []
        wb_obj = None
        sheet = None
        for file in os.listdir(utils.DIR_OUTPUT_):
            if 'ExportAppuisCommunBrut' in file:
                print(file)
                wb_obj = openpyxl.load_workbook(utils.DIR_OUTPUT_ + file)


        for y, sheetname in enumerate(wb_obj.sheetnames):
            if 'Export' in sheetname:
                sheet = wb_obj.worksheets[y]

        for col in range(1, 14):
            list_field_etude_export_brut.append(sheet.cell(column=col, row=1).value)

        for row in range(2, sheet.max_row):
            dict_etude_export_brut = {}
            for i, col in enumerate(range(1, 14)):
                dict_etude_export_brut[list_field_etude_export_brut[i]] = sheet.cell(column=col, row=row).value
            list_etudes_export_brut.append(dict_etude_export_brut)


        for row in list_etudes_export_brut:
            for i, etude in enumerate(self.data):
                if row['Numéro affaire Opérateur'] == etude["Nom d'affaire E-PLAN"]:
                    if row["Etat de l''affaire"] == 'Etude Validée':
                        print('validé', row["Numéro affaire Enedis"])
                        self.data[i]['numéro_affaire_enedis'] = row["Numéro affaire Enedis"]
                else:
                    adresse_etude = re.sub('[;,-_°.\'\"/ ]', '', etude["Adresse"])
                    adresse_etude = re.sub('[éèê]', 'e', adresse_etude)
                    adresse_etude = adresse_etude.upper()
                    adresse_export = re.sub('[;,-_°.\'\"/ ]', '', row['Localisation du projet'])
                    adresse_export = re.sub('[éèê]', 'e', adresse_export)
                    adresse_export = adresse_export.upper()

                    if etude["Numéro PMZ GEOFIBRE"] in row['Numéro affaire Opérateur'] and etude["N° Plan"] + '-' in row['Numéro affaire Opérateur'] and adresse_etude == adresse_export:
                        print(row['Numéro affaire Opérateur'], row['Numéro affaire Opérateur'])
                        if row["Etat de l''affaire"] == 'Etude Validée':
                            print('validé', row["Numéro affaire Enedis"])
                            self.data[i]['numéro_affaire_enedis'] = row["Numéro affaire Enedis"]





    def iterate_over_etudes(self):
        driver = self.driver

        for row in self.data:



            writer = create_doc.writeDoc(row,self.operateur,self.date,self.departement)
            writer.replace_text()
            pdf = writer.save_doc()


                #driver.get("https://www.e-plans.fr/DemandeAppuisCommuns/List")
                #driver.find_element_by_xpath('//input[@id="numAffaire"]').send_keys(row['numéro_affaire_enedis'])
                #driver.find_element_by_xpath('/html/body/nav/div/div[2]/form/div/span/input').click()

                #WebDriverWait(driver, timeout=30).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="etapes-list"]/li[1]/ul/li[4]/a'))).click()
                #sleep(2)
                #driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/div[1]/div/div[1]/div/div[1]/div[4]/form[1]/div/fieldset/table/tbody/tr/td/div/div/div/div[2]/input").send_keys(pdf)





        #print(self.data)
